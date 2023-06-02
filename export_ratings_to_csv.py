import os
import openpyxl
import django
from django.core.exceptions import ObjectDoesNotExist

# Set up Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'peerreview.settings')
django.setup()

from peerapp.models import Rating

# Open the Excel file for reading and writing
filepath = r'C:/Users/madhu/Desktop/Peer Review/peerreview/project final details_updated.xlsx'
workbook = openpyxl.load_workbook(filepath)
sheet = workbook.active

# Iterate over the rows starting from row 2
for row in sheet.iter_rows(min_row=2):
    usn = row[2].value  # Assuming 'USN' is in column 3 (index 2)

    try:
        # Query the Rating model based on the USN
        rating = Rating.objects.get(usn=usn)

        # Update the average_rating value in the Excel row
        sheet.cell(row=row[0].row, column=4).value = rating.average_rating
        # Assuming 'AVERAGE RATING' is in column 4

    except ObjectDoesNotExist:
        # Handle the case where the USN is not found in the database
        pass

# Save the updated workbook to a new file
output_filepath = r'C:/Users/madhu/Desktop/Peer Review/peerreview/project final details_updated.xlsx'
workbook.save(output_filepath)
workbook.close()
